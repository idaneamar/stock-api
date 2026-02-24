#!/usr/bin/env python3
"""
Summary_optsp.py (Combined: technical + plain-English report in ONE Excel)

Output (one file):
- backtest_optsp_summaries/Summary_optsp.xlsx

Tabs for non-technical readers:
- EXECUTIVE_SUMMARY
- KEY_METRICS
- YEARLY_COMPARISON
- STRATEGY_EXPLANATION

Technical tabs:
- Overview, Daily, Virtual_by_year
- P&L, P&L_by_year, P&L_detail_sample
- P&L_TPSL, P&L_TPSL_by_year, P&L_TPSL_detail_sample

Run:
  caffeinate -i python3 backtest_optsp_system.py --years 2024 2025 --step virtual --analysis-type full
  caffeinate -i python3 Summary_optsp.py
"""

from __future__ import annotations

import os
import argparse
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = os.getcwd()
VIRTUAL_DIR = os.path.join(BASE_DIR, "backtest_optsp_virtual")
SUMMARIES_DIR = os.path.join(BASE_DIR, "backtest_optsp_summaries")
HIST_DIR = os.path.join(BASE_DIR, "historical_data_cache")

TP_PCT = float(os.environ.get("OPTSP_TP_PCT", "0.50"))
SL_MULT = float(os.environ.get("OPTSP_SL_MULT", "2.00"))
TP_HOLD_DAYS = int(os.environ.get("OPTSP_TP_HOLD_DAYS", "14"))
EXIT_DAYS_BEFORE_EXP = int(os.environ.get("OPTSP_EXIT_DAYS_BEFORE_EXP", "0"))

DETAIL_SAMPLE_ROWS = 2000

_price_cache: Dict[str, pd.DataFrame] = {}

def _load_virtual_year(year: int) -> Optional[pd.DataFrame]:
    p = os.path.join(VIRTUAL_DIR, f"virtual_{year}.csv")
    if not os.path.exists(p):
        return None
    df = pd.read_csv(p)
    if df.empty:
        return None
    df["run_date"] = pd.to_datetime(df["run_date"], errors="coerce")
    df["exp"] = pd.to_datetime(df["exp"], errors="coerce")
    df = df.dropna(subset=["run_date", "exp"]).copy()
    df["year"] = df["run_date"].dt.year
    return df

def _read_price_df(ticker: str) -> Optional[pd.DataFrame]:
    t = str(ticker).upper().replace(".US", "").strip()
    if t in _price_cache:
        df = _price_cache[t]
        return None if df.empty else df

    candidates = [
        os.path.join(HIST_DIR, f"{t}.parquet"),
        os.path.join(HIST_DIR, f"{t}.US.parquet"),
    ]
    p = next((x for x in candidates if os.path.exists(x)), None)
    if not p:
        _price_cache[t] = pd.DataFrame()
        return None

    try:
        df = pd.read_parquet(p)
    except Exception:
        _price_cache[t] = pd.DataFrame()
        return None

    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df = df.dropna(subset=["date"]).set_index("date")
    else:
        df.index = pd.to_datetime(df.index, errors="coerce")
        df = df[~df.index.isna()]
    df = df.sort_index()
    _price_cache[t] = df
    return df

def _get_ohlc(df: pd.DataFrame, d: pd.Timestamp) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    if df is None or df.empty:
        return None, None, None
    d = pd.to_datetime(d).normalize()

    try:
        row = df.loc[d]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[-1]
    except Exception:
        idx = df.index[df.index <= d]
        if len(idx) == 0:
            return None, None, None
        row = df.loc[idx[-1]]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[-1]

    def pick(*names):
        for n in names:
            if n in row.index and pd.notna(row[n]):
                return float(row[n])
        return None

    close = pick("close", "Close", "adj_close", "Adj Close", "adjusted_close")
    low = pick("low", "Low")
    high = pick("high", "High")
    if close is None:
        return None, None, None
    if low is None:
        low = close
    if high is None:
        high = close
    return low, high, close

def _compute_intrinsic_pnl(spot_at_exp: float, short_put: float, long_put: float, short_call: float, long_call: float, credit: float) -> float:
    put_loss = 0.0
    if spot_at_exp < short_put:
        put_loss = min(short_put - spot_at_exp, short_put - long_put)
    call_loss = 0.0
    if spot_at_exp > short_call:
        call_loss = min(spot_at_exp - short_call, long_call - short_call)
    return float(credit) - float(put_loss + call_loss)

@dataclass
class TPSLResult:
    pnl_per_share: float
    exit_reason: str
    exit_date: pd.Timestamp
    spot_exit: float

def _simulate_tpsl(row: pd.Series, px: pd.DataFrame) -> Optional[TPSLResult]:
    try:
        run_date = pd.to_datetime(row["run_date"]).normalize()
        exp = pd.to_datetime(row["exp"]).normalize()

        exp_effective = exp
        if EXIT_DAYS_BEFORE_EXP > 0:
            tmp = exp - pd.Timedelta(days=int(EXIT_DAYS_BEFORE_EXP))
            if tmp > run_date:
                exp_effective = tmp

        credit = float(row["net_credit"])
        width = float(row.get("width", abs(float(row["short_call"]) - float(row["short_put"])) ))
        max_loss = float(width) - float(credit)
        if max_loss <= 0:
            max_loss = float(width)

        short_put = float(row["short_put"])
        short_call = float(row["short_call"])
        long_put = float(row["long_put"])
        long_call = float(row["long_call"])

        tp_date = run_date + pd.Timedelta(days=int(TP_HOLD_DAYS))
        if tp_date > exp_effective:
            tp_date = exp_effective

        d = run_date
        while d <= exp_effective:
            low, high, close = _get_ohlc(px, d)
            if close is None:
                d += pd.Timedelta(days=1)
                continue

            if (low is not None and low <= short_put) or (high is not None and high >= short_call):
                pnl_sl = max(credit - SL_MULT * credit, -max_loss)
                return TPSLResult(float(pnl_sl), "sl_touch", d, float(close))

            if d >= tp_date and tp_date < exp_effective:
                pnl_tp = TP_PCT * credit
                return TPSLResult(float(pnl_tp), "tp", d, float(close))

            d += pd.Timedelta(days=1)

        _, _, close_exp = _get_ohlc(px, exp_effective)
        if close_exp is None:
            return None
        pnl_exp = _compute_intrinsic_pnl(float(close_exp), short_put, long_put, short_call, long_call, credit)
        return TPSLResult(float(pnl_exp), "expiry", exp_effective, float(close_exp))
    except Exception:
        return None

def _pnl_summary(df_all: pd.DataFrame, pnl_df: pd.DataFrame) -> pd.DataFrame:
    if pnl_df.empty:
        return pd.DataFrame([{
            "total_recommendations": int(len(df_all)),
            "with_underlying_price": 0,
            "missing_price_or_exp": int(len(df_all)),
            "total_pnl_per_share": 0.0,
            "total_pnl_per_contract": 0.0,
            "wins": 0, "losses": 0, "flats": 0,
            "win_rate_pct": 0.0,
        }])
    wins = int((pnl_df["pnl_per_share"] > 0).sum())
    losses = int((pnl_df["pnl_per_share"] < 0).sum())
    flats = int((pnl_df["pnl_per_share"] == 0).sum())
    denom = wins + losses + flats
    win_rate = (wins / denom * 100.0) if denom else 0.0
    return pd.DataFrame([{
        "total_recommendations": int(len(df_all)),
        "with_underlying_price": int(len(pnl_df)),
        "missing_price_or_exp": int(len(df_all) - len(pnl_df)),
        "total_pnl_per_share": float(pnl_df["pnl_per_share"].sum()),
        "total_pnl_per_contract": float(pnl_df["pnl_per_contract"].sum()),
        "wins": wins, "losses": losses, "flats": flats,
        "win_rate_pct": float(win_rate),
    }])

def _style_sheet(ws, freeze="A2", header_row=1, col_widths: Optional[Dict[int, float]] = None) -> None:
    ws.freeze_panes = freeze
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    if col_widths:
        for col_idx, w in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w
    else:
        for col in range(1, ws.max_column + 1):
            mx = 10
            for row in range(1, min(ws.max_row, 200) + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                mx = max(mx, min(60, len(str(v)) + 2))
            ws.column_dimensions[get_column_letter(col)].width = mx

def _add_pretty_sheets(out_path: str, overview_df: pd.DataFrame, pnl_total: pd.DataFrame, tpsl_total: pd.DataFrame, tpsl_by_year: pd.DataFrame) -> None:
    wb = load_workbook(out_path)

    for name in ["EXECUTIVE_SUMMARY", "KEY_METRICS", "YEARLY_COMPARISON", "STRATEGY_EXPLANATION"]:
        if name in wb.sheetnames:
            wb.remove(wb[name])

    ws_exec = wb.create_sheet("EXECUTIVE_SUMMARY", 0)
    ws_kpi = wb.create_sheet("KEY_METRICS", 1)
    ws_year = wb.create_sheet("YEARLY_COMPARISON", 2)
    ws_exp = wb.create_sheet("STRATEGY_EXPLANATION", 3)

    tpsl = tpsl_total.iloc[0].to_dict() if not tpsl_total.empty else {}
    base = pnl_total.iloc[0].to_dict() if not pnl_total.empty else {}

    years_list = [str(int(x)) for x in overview_df["year"].dropna().tolist()] if (not overview_df.empty and "year" in overview_df.columns) else []
    years_txt = ", ".join(years_list) if years_list else "N/A"

    total_trades = int(tpsl.get("with_underlying_price", tpsl.get("total_recommendations", 0)) or 0)
    win_rate = float(tpsl.get("win_rate_pct", 0.0) or 0.0)
    total_pnl = float(tpsl.get("total_pnl_per_contract", 0.0) or 0.0)
    avg_per_trade = (total_pnl / total_trades) if total_trades else 0.0

    ws_exec["A1"] = "סיכום ביצועי המערכת (בשפה פשוטה)"
    ws_exec["A3"] = "מה נבדק:"
    ws_exec["B3"] = f"אסטרטגיית Iron Condor, עד 10 עסקאות ביום, על נתונים היסטוריים בשנים: {years_txt}."
    ws_exec["A5"] = "מה המסקנה:"
    ws_exec["B5"] = (
        "כאשר משתמשים בניהול סיכונים (Take Profit ו-Stop Loss), מתקבלת תמונת ביצועים יציבה יותר. "
        "בלי ניהול סיכונים, הפסדים נדירים אך גדולים יכולים למחוק הרבה רווחים קטנים."
    )
    ws_exec["A7"] = "המספרים המרכזיים (עם ניהול סיכונים):"
    ws_exec["A8"] = "מספר עסקאות שנבדקו:"
    ws_exec["B8"] = total_trades
    ws_exec["A9"] = "אחוז הצלחה (Win Rate):"
    ws_exec["B9"] = round(win_rate, 2)
    ws_exec["A10"] = "רווח/הפסד כולל ($):"
    ws_exec["B10"] = round(total_pnl, 2)
    ws_exec["A11"] = "רווח ממוצע לעסקה ($):"
    ws_exec["B11"] = round(avg_per_trade, 2)
    ws_exec["A13"] = "הערה חשובה:"
    ws_exec["B13"] = (
        "זו סימולציה היסטורית ולא הבטחה לעתיד. "
        "במסחר אמיתי יש עמלות, החלקה (slippage), סיכוני ביצוע ומרווחי BID/ASK."
    )

    ws_exec["A1"].font = Font(bold=True, size=16)
    for a in ["A3", "A5", "A7", "A13"]:
        ws_exec[a].font = Font(bold=True)

    ws_exec.column_dimensions["A"].width = 28
    ws_exec.column_dimensions["B"].width = 95
    ws_exec["B5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_exec["B13"].alignment = Alignment(wrap_text=True, vertical="top")

    ws_kpi.append(["מדד", "ערך"])
    ws_kpi.append(["שנים שנבדקו", years_txt])
    ws_kpi.append(["מספר עסקאות (עם מחיר/נתונים זמינים)", total_trades])
    ws_kpi.append(["אחוז הצלחה (Win Rate %)", round(win_rate, 2)])
    ws_kpi.append(["רווח/הפסד כולל ($)", round(total_pnl, 2)])
    ws_kpi.append(["רווח ממוצע לעסקה ($)", round(avg_per_trade, 2)])

    base_trades = int(base.get("with_underlying_price", 0) or 0)
    base_pnl = float(base.get("total_pnl_per_contract", 0.0) or 0.0)
    base_wr = float(base.get("win_rate_pct", 0.0) or 0.0)
    ws_kpi.append(["---", "---"])
    ws_kpi.append(["(להשוואה) ללא ניהול סיכונים (Hold עד פקיעה) – Win Rate %", round(base_wr, 2)])
    ws_kpi.append(["(להשוואה) ללא ניהול סיכונים (Hold עד פקיעה) – P&L כולל ($)", round(base_pnl, 2)])
    ws_kpi.append(["(להשוואה) מספר עסקאות זמינות לחישוב", base_trades])
    _style_sheet(ws_kpi, col_widths={1: 52, 2: 32})

    ws_year.append(["שנה", "מספר עסקאות", "רווח/הפסד כולל ($)", "ניצחונות", "הפסדים"])
    if tpsl_by_year is not None and not tpsl_by_year.empty:
        for _, r in tpsl_by_year.iterrows():
            ws_year.append([int(r["year"]), int(r["recs"]), float(r["total_pnl_per_contract"]), int(r["wins"]), int(r["losses"])])
    _style_sheet(ws_year, col_widths={1: 10, 2: 16, 3: 18, 4: 12, 5: 12})

    ws_exp["A1"] = "הסבר קצר על האסטרטגיה (למי שלא מכיר)"
    ws_exp["A3"] = (
        "המערכת מחפשת עסקאות אופציות מסוג Iron Condor על מניות גדולות וסחירות. "
        "העסקה מרוויחה אם מחיר המניה נשאר בתוך טווח (בין הסטרייקים) עד היציאה מהעסקה."
    )
    ws_exp["A5"] = (
        "ניהול סיכונים (בסימולציה):\n"
        f"- Take Profit: סגירה ברווח של ~{int(TP_PCT*100)}% מהפרמיה.\n"
        f"- Stop Loss: סגירה בהפסד של ~{SL_MULT:.2f}× מהפרמיה אם יש 'Touch' לסטרייק.\n"
        "- כלל חשיפה: מקסימום 10 עסקאות ביום.\n"
    )
    ws_exp["A8"] = (
        "מה חשוב לדעת:\n"
        "1) זו סימולציה היסטורית – לא הבטחה.\n"
        "2) במציאות יש עמלות, מרווחי BID/ASK והחלקה.\n"
        "3) המטרה כאן היא להעריך יציבות, Win Rate ורווחיות לאורך זמן."
    )
    ws_exp["A1"].font = Font(bold=True, size=14)
    for cell in ["A3", "A5", "A8"]:
        ws_exp[cell].alignment = Alignment(wrap_text=True, vertical="top")
    ws_exp.column_dimensions["A"].width = 120
    wb.save(out_path)

def build_summary_excel(years: List[int]) -> str:
    os.makedirs(SUMMARIES_DIR, exist_ok=True)
    out_path = os.path.join(SUMMARIES_DIR, "Summary_optsp.xlsx")

    frames = [df for y in years if (df := _load_virtual_year(y)) is not None and not df.empty]
    if not frames:
        raise RuntimeError("No virtual_*.csv files found. Run virtual step first.")

    df_all = pd.concat(frames, ignore_index=True)
    df_all["year"] = df_all["run_date"].dt.year

    overview = df_all.groupby("year").agg(
        days=("run_date", lambda x: int(x.dt.normalize().nunique())),
        total_recs=("run_date", "count"),
        avg_pop=("pop_est", "mean"),
        avg_credit=("net_credit", "mean"),
    ).reset_index()

    daily = df_all.groupby(df_all["run_date"].dt.date).agg(
        year=("year", "first"),
        count=("run_date", "count"),
        avg_pop=("pop_est", "mean"),
        avg_credit=("net_credit", "mean"),
    ).reset_index().rename(columns={"run_date": "date"})
    daily["date"] = pd.to_datetime(daily["date"])

    virtual_by_year = df_all.groupby("year").agg(
        virtual_recs=("run_date", "count"),
        avg_pop_est=("pop_est", "mean"),
        avg_net_credit=("net_credit", "mean"),
    ).reset_index()

    pnl_rows, detail_rows = [], []
    for _, r in df_all.iterrows():
        ticker = str(r.get("ticker", "")).upper().replace(".US", "").strip()
        px = _read_price_df(ticker)
        if px is None or px.empty:
            continue
        exp = pd.to_datetime(r["exp"]).normalize()
        _, _, close_exp = _get_ohlc(px, exp)
        if close_exp is None:
            continue
        pnl_ps = _compute_intrinsic_pnl(float(close_exp), float(r["short_put"]), float(r["long_put"]), float(r["short_call"]), float(r["long_call"]), float(r["net_credit"]))
        out = {
            "run_date": pd.to_datetime(r["run_date"]).date().isoformat(),
            "ticker": ticker,
            "exp": exp.date().isoformat(),
            "net_credit": float(r["net_credit"]),
            "spot_at_exp": float(close_exp),
            "pnl_per_share": float(pnl_ps),
            "pnl_per_contract": float(pnl_ps) * 100.0,
            "outcome": "win" if pnl_ps > 0 else ("loss" if pnl_ps < 0 else "flat"),
            "year": int(r["year"]),
        }
        pnl_rows.append(out)
        if len(detail_rows) < DETAIL_SAMPLE_ROWS:
            detail_rows.append(out)
    pnl_df = pd.DataFrame(pnl_rows)
    pnl_total = _pnl_summary(df_all, pnl_df)
    pnl_by_year = pnl_df.groupby("year").agg(
        recs=("pnl_per_contract", "count"),
        total_pnl_per_share=("pnl_per_share", "sum"),
        total_pnl_per_contract=("pnl_per_contract", "sum"),
        wins=("pnl_per_share", lambda x: int((x > 0).sum())),
        losses=("pnl_per_share", lambda x: int((x < 0).sum())),
    ).reset_index() if not pnl_df.empty else pd.DataFrame()

    tpsl_rows, tpsl_detail_rows = [], []
    for _, r in df_all.iterrows():
        ticker = str(r.get("ticker", "")).upper().replace(".US", "").strip()
        px = _read_price_df(ticker)
        if px is None or px.empty:
            continue
        res = _simulate_tpsl(r, px)
        if res is None:
            continue
        trow = {
            "run_date": pd.to_datetime(r["run_date"]).date().isoformat(),
            "ticker": ticker,
            "exp": pd.to_datetime(r["exp"]).date().isoformat(),
            "net_credit": float(r["net_credit"]),
            "exit_reason": res.exit_reason,
            "exit_date": pd.to_datetime(res.exit_date).date().isoformat(),
            "spot_exit": float(res.spot_exit),
            "pnl_per_share": float(res.pnl_per_share),
            "pnl_per_contract": float(res.pnl_per_share) * 100.0,
            "year": int(r["year"]),
        }
        tpsl_rows.append(trow)
        if len(tpsl_detail_rows) < DETAIL_SAMPLE_ROWS:
            tpsl_detail_rows.append(trow)
    tpsl_df = pd.DataFrame(tpsl_rows)
    tpsl_total = _pnl_summary(df_all, tpsl_df)
    tpsl_by_year = tpsl_df.groupby("year").agg(
        recs=("pnl_per_contract", "count"),
        total_pnl_per_share=("pnl_per_share", "sum"),
        total_pnl_per_contract=("pnl_per_contract", "sum"),
        wins=("pnl_per_share", lambda x: int((x > 0).sum())),
        losses=("pnl_per_share", lambda x: int((x < 0).sum())),
    ).reset_index() if not tpsl_df.empty else pd.DataFrame()

    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        overview.to_excel(w, index=False, sheet_name="Overview")
        daily.to_excel(w, index=False, sheet_name="Daily")
        virtual_by_year.to_excel(w, index=False, sheet_name="Virtual_by_year")
        pnl_total.to_excel(w, index=False, sheet_name="P&L")
        pnl_by_year.to_excel(w, index=False, sheet_name="P&L_by_year")
        pd.DataFrame(detail_rows).to_excel(w, index=False, sheet_name="P&L_detail_sample")
        tpsl_total.to_excel(w, index=False, sheet_name="P&L_TPSL")
        tpsl_by_year.to_excel(w, index=False, sheet_name="P&L_TPSL_by_year")
        pd.DataFrame(tpsl_detail_rows).to_excel(w, index=False, sheet_name="P&L_TPSL_detail_sample")

    _add_pretty_sheets(out_path, overview, pnl_total, tpsl_total, tpsl_by_year)
    return out_path

def main() -> None:
    global TP_PCT, SL_MULT, TP_HOLD_DAYS, EXIT_DAYS_BEFORE_EXP
    ap = argparse.ArgumentParser()
    ap.add_argument("--years", nargs="+", type=int, default=None)
    ap.add_argument("--tp-pct", type=float, default=TP_PCT)
    ap.add_argument("--sl-mult", type=float, default=SL_MULT)
    ap.add_argument("--tp-hold-days", type=int, default=TP_HOLD_DAYS)
    ap.add_argument("--exit-days-before-exp", type=int, default=EXIT_DAYS_BEFORE_EXP)
    args = ap.parse_args()

    TP_PCT = float(args.tp_pct)
    SL_MULT = float(args.sl_mult)
    TP_HOLD_DAYS = int(args.tp_hold_days)
    EXIT_DAYS_BEFORE_EXP = int(args.exit_days_before_exp)

    years = args.years
    if not years:
        years = []
        if os.path.isdir(VIRTUAL_DIR):
            for fn in os.listdir(VIRTUAL_DIR):
                if fn.startswith("virtual_") and fn.endswith(".csv"):
                    try:
                        years.append(int(fn.replace("virtual_", "").replace(".csv", "")))
                    except Exception:
                        pass
        years = sorted(set(years))

    if not years:
        raise RuntimeError("No years provided and none inferred. Run virtual step first.")

    out = build_summary_excel(years)
    print(f"Saved: {out}")

if __name__ == "__main__":
    main()
